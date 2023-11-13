import openpyxl
import os
import xlwings as xlw
import tkinter as tk


wb = xlw.books.active
ws = wb.sheets["Tong_hop"]
# Tạo cửa sổ làm việc
window = tk.Tk()
window.geometry("200x250")
# Khu vực chứa function
def task():
    rid = nametb.get()
    if "sơ nhân thân" in str(ws["Y"+str(rid)].value):
        clone = str(ws["Y"+str(rid)].value).split("\n")
        for text in clone:
            if "sơ nhân thân" in text:
                clone.remove(text)
        ws["Y"+str(rid)].value = "\n".join(clone)
    clone = str(ws["Z"+str(rid)].value).split("\n")
    clone.append("Hồ sơ nhân thân sao y")
    ws["Z" + str(rid)].value = "\n".join(clone)
# Tạo label và ô nhập liệu (entry Box)
namelb = tk.Label(window, text= "Số hàng: ")
nametb = tk.Entry(window)
rid = nametb.get()
# Tạo nút
button = tk.Button(window, text = "Chạy", command = task)
# Định đạng bảng làm việc
namelb.grid(row = 1 , column= 0)
nametb.grid(row = 1, column = 1)
nametb.focus()
button.grid(row=2, columnspan= 2)

#Xuất bảng
window.mainloop()