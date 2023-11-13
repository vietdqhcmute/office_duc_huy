from libs.Functions import *
def Mail_Merged_Mau04(wb_path, ws_name, doc_sample_path, des_folder):
    import openpyxl, docx, shutil
    from os.path import join
#1. INPUT dư liệu excel
    wb = openpyxl.load_workbook(wb_path, data_only= True)
    ws = wb[ws_name]
#2. Tag «» Mail_merge cho dòng tên cột (dòng 1)
    dict_cl_name = {}
    count_cl_name = 1
    for cl_idx in range(1, ws.max_column+1):
        cl_name = icell(ws, 1, cl_idx)
        name_tag = "«"+ cl_name + "»"
        w_icell(ws, 1, cl_idx, name_tag)
        dict_cl_name[name_tag] = count_cl_name
        count_cl_name += 1
#3. Duyệt từng dòng của file excel
    for r_idx in range(2, ws.max_row + 1):
####. Trường hợp này đặc biệt vì chỉ tạo Merged khi có số MBV
        mbv = icell(ws, r_idx, 2)
        if mbv != "None":
#4. Tạo và đặt tên file word kết quả
####. Chỉnh tên muốn đặt
            out_name = icell(ws, r_idx, 2) + "_" + mbv + ".docx"
            shutil.copy(doc_sample_path, join(des_folder, out_name))
#5. Mở từng file vừa tạo để gắn giá trị theo Tag Mail_Merge
            doc = docx.Document(join(des_folder, out_name))
####. Duyệt text
            for para in doc.paragraphs:
                for run in para.runs:
                    if run.text in dict_cl_name.keys():
                        content = icell(ws, r_idx, dict_cl_name[run.text])
                        if content != "None":
                            run.text = content
                        else:
                            run.text = ""
####. Duyệt Table
            for tab in doc.tables:
                for row in tab.rows:
                    for cell in row.cells:
                        for para in cell.paragraphs:
                            for run in para.runs:
                                if run.text in dict_cl_name.keys():
                                    content = icell(ws, r_idx, dict_cl_name[run.text])
                                    if content != "None":
                                        run.text = content
                                    else:
                                        run.text = "" 
            doc.save(join(des_folder,out_name))