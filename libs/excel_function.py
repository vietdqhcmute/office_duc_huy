import os
import openpyxl

# Functions related to sheet
def get_sheet_by_index(excel_path, sheet_index):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[sheet_index]]
    return ws

def get_sheet_by_name(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    return ws

# Functions to read cells
def read_cell_by_letter(ws, row_index, column_letter):
    value = str(ws[str(column_letter) + str(row_index)].value)
    return value

def read_cell_by_index(ws, row_index, column_index):
    value = str(ws.cell(row = row_index, column = column_index).value)
    return value

def get_cell_coordinate(ws, row_index, column_index):
    coordinate = ws.cell(row = row_index, column = column_index).coordinate
    return coordinate

# Functions to write to cells
def write_cell_by_index(ws, row_index, column_index, content):
    ws.cell(row = row_index, column = column_index).value = content

def write_cell_by_letter(ws, row_index, column_letter, content):
    ws[str(column_letter) + str(row_index)] = content
